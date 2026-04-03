"""
MT5 P&L Studio – Manager API Edition
──────────────────────────────────────
Two report modes:

  Deal P&L    – PnL per trade = Profit + Commission + Swap
  Equity P&L  – Net P&L = Closing Equity − Opening Equity − Net D/W − Net Credit − Bonus
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from io import BytesIO

from mt5_connector import (
    connect_mt5,
    disconnect_mt5,
    get_deals,
    get_users,
    get_user_registrations,
    deals_to_dicts,
    get_daily_reports,
    daily_to_dicts,
)
from calculations import (
    build_deals_dataframe,
    compute_client_report,
    compute_group_summary,
    compute_symbol_summary,
    get_top_gainers,
    get_top_losers,
    compute_kpis,
    compute_equity_report,
    compute_equity_group_summary,
    compute_equity_kpis,
    compute_account_category_summary,
)
from db import (
    init_db,
    verify_user,
    change_password,
    get_mt5_profiles,
    get_mt5_profile_decrypted,
    save_mt5_profile,
    delete_mt5_profile,
    list_users,
    create_user,
    delete_user,
)

# ═══════════════════════════════════════════════════════════
# PAGE CONFIG & GLOBAL CSS
# ═══════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Reporting Tool",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# PWA meta tags — makes the app installable on mobile
st.markdown("""
<link rel="manifest" href="/app/static/manifest.json">
<meta name="theme-color" content="#3b82f6">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="Reporting Tool">
<link rel="apple-touch-icon" href="/app/static/icon-192.png">
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════
# AUTH GATE
# ═══════════════════════════════════════════════════════════
init_db()

if "authenticated" not in st.session_state:
    st.session_state.update(
        authenticated=False,
        user_id=None,
        username=None,
        is_admin=False,
        must_change_pw=False,
    )


def _render_login_screen():
    st.markdown(
        "<style>.block-container{max-width:440px!important;padding-top:6rem!important;}</style>",
        unsafe_allow_html=True,
    )
    st.markdown("## Reporting Tool")
    st.markdown("##### Sign in to continue")
    with st.form("login_form"):
        uname = st.text_input("Username")
        pwd   = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Sign In", use_container_width=True)
    if submitted:
        user = verify_user(uname, pwd)
        if user:
            st.session_state.update(
                authenticated=True,
                user_id=user["id"],
                username=user["username"],
                is_admin=bool(user["is_admin"]),
                must_change_pw=bool(user["must_change_pw"]),
            )
            st.rerun()
        else:
            st.error("Invalid username or password.")
    st.stop()


def _render_change_password_screen():
    st.markdown("## Change Password Required")
    st.info("You must set a new password before continuing.")
    with st.form("change_pw_form"):
        new_pw  = st.text_input("New Password", type="password")
        confirm = st.text_input("Confirm Password", type="password")
        submitted = st.form_submit_button("Set Password", use_container_width=True)
    if submitted:
        if len(new_pw) < 8:
            st.error("Password must be at least 8 characters.")
        elif new_pw != confirm:
            st.error("Passwords do not match.")
        else:
            change_password(st.session_state.user_id, new_pw)
            st.session_state.must_change_pw = False
            st.rerun()
    st.stop()


if not st.session_state.authenticated:
    _render_login_screen()

if st.session_state.must_change_pw:
    _render_change_password_screen()

st.markdown(
    """
<style>
/* ── Base ── */
body, .main, [data-testid="stAppViewContainer"]{
  background: #f0f4f8 !important;
  color: #0f172a;
  font-family: ui-sans-serif, system-ui, -apple-system, "Segoe UI", Roboto, Arial, sans-serif;
}
.block-container{max-width:1500px; padding-top:1.2rem; padding-bottom:3rem;}
h1,h2,h3,h4{color:#0f172a;}

/* ── Hero banner ── */
.hero{
  border-radius:20px;
  padding:1.4rem 1.8rem;
  background: linear-gradient(135deg, #0b1220 0%, #1e293b 55%, #0b1220 100%);
  color:#f8fafc;
  border: 1px solid rgba(255,255,255,0.10);
  box-shadow: 0 16px 48px rgba(15,23,42,0.30);
  position:relative; overflow:hidden; margin-bottom:1rem;
}
.hero:before{
  content:""; position:absolute; inset:-100px -100px auto auto;
  width:220px; height:220px;
  background: radial-gradient(circle, rgba(99,102,241,.40), transparent);
}
.hero:after{
  content:""; position:absolute; inset:auto auto -120px -120px;
  width:260px; height:260px;
  background: radial-gradient(circle, rgba(16,185,129,.30), transparent);
}
.hero-badge{
  display:inline-flex; align-items:center; gap:.5rem;
  padding:.22rem .75rem; border-radius:999px; font-size:.78rem;
  color:rgba(248,250,252,.85);
  border:1px solid rgba(255,255,255,.25);
  background:rgba(255,255,255,.08);
  margin-bottom:.6rem;
}
.hero-title{margin:.5rem 0 .25rem 0; font-size:1.9rem; font-weight:800; line-height:1.2;}
.hero-sub{margin:0; color:rgba(248,250,252,.75); max-width:900px; line-height:1.6; font-size:.95rem;}

/* ── Section card ── */
.section{
  margin-top:1rem; padding:1.1rem 1.2rem;
  background:#ffffff;
  border:1px solid #e2e8f0;
  border-radius:16px;
  box-shadow: 0 4px 16px rgba(2,6,23,0.06);
}
.section-title{display:flex; align-items:center; justify-content:space-between; gap:1rem; margin-bottom:.5rem;}
.section-title h2{margin:0; font-size:1.1rem; font-weight:700; color:#0f172a;}
.pill{
  display:inline-flex; align-items:center; gap:.4rem;
  padding:.2rem .7rem; border-radius:999px;
  background:rgba(99,102,241,0.10);
  border:1px solid rgba(99,102,241,0.25);
  color:#4338ca; font-size:.78rem; font-weight:600;
}
.pill-green{
  display:inline-flex; align-items:center; gap:.4rem;
  padding:.2rem .7rem; border-radius:999px;
  background:rgba(5,150,105,0.10);
  border:1px solid rgba(5,150,105,0.25);
  color:#059669; font-size:.78rem; font-weight:600;
}

/* ── KPI Metric card ── */
.metric{
  background:#ffffff;
  border:1px solid #e2e8f0;
  border-radius:14px; padding:1rem 1.1rem;
  box-shadow: 0 4px 14px rgba(2,6,23,0.06);
}
.metric .k{font-size:.7rem; letter-spacing:.10em; text-transform:uppercase; color:#64748b; font-weight:600;}
.metric .v{font-size:1.45rem; font-weight:800; color:#0f172a; margin-top:.25rem; line-height:1.2;}
.metric .s{font-size:.76rem; color:#94a3b8; margin-top:.25rem;}
.metric .v.green{color:#059669 !important;}
.metric .v.red{color:#dc2626 !important;}
.metric .v.blue{color:#2563eb !important;}


/* ── Sidebar ── */
[data-testid="stSidebar"]{
  background: linear-gradient(180deg, #0b1220 0%, #111827 100%) !important;
  border-right: 1px solid rgba(255,255,255,0.07) !important;
}
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div,
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] small{
  color: #f1f5f9 !important;
}
[data-testid="stSidebar"] .stTextInput > div > div > input{
  background: #ffffff !important;
  border: 1px solid #cbd5e1 !important;
  color: #0f172a !important;
  border-radius: 10px !important;
  caret-color: #0f172a !important;
}
[data-testid="stSidebar"] .stTextInput > div > div > input::placeholder{
  color: #94a3b8 !important;
}
[data-testid="stSidebar"] .stNumberInput input{
  background: #ffffff !important;
  border: 1px solid #cbd5e1 !important;
  color: #0f172a !important;
  border-radius: 10px !important;
  caret-color: #0f172a !important;
}
[data-testid="stSidebar"] .stDateInput input{
  background: #ffffff !important;
  border: 1px solid #cbd5e1 !important;
  color: #0f172a !important;
  border-radius: 10px !important;
  caret-color: #0f172a !important;
}
[data-testid="stSidebar"] .stSlider [data-baseweb="slider"]{
  background: rgba(255,255,255,0.15) !important;
}
[data-testid="stSidebar"] .stButton > button{
  background: linear-gradient(135deg, #6366f1, #4f46e5) !important;
  color: #ffffff !important;
  border: none !important;
  border-radius: 12px !important;
  font-weight: 700 !important;
  padding: .6rem 1rem !important;
  box-shadow: 0 4px 14px rgba(99,102,241,.40) !important;
}
[data-testid="stSidebar"] .stButton > button:hover{
  background: linear-gradient(135deg, #4f46e5, #4338ca) !important;
}
[data-testid="stSidebar"] hr{border-color: rgba(255,255,255,0.12) !important;}

/* ── File uploader dropzone: white bg, dark text ── */
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"]{
  background: #ffffff !important;
  border: 1px solid #cbd5e1 !important;
  border-radius: 10px !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] span,
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] p,
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] div,
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] small,
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button{
  color: #0f172a !important;
}
/* ── Uploaded filename row: keep text white ── */
[data-testid="stSidebar"] [data-testid="stFileUploader"] [data-testid="stFileUploaderFileName"],
[data-testid="stSidebar"] [data-testid="stFileUploader"] [class*="uploadedFile"] span,
[data-testid="stSidebar"] [data-testid="stFileUploader"] [class*="uploadedFile"] div,
[data-testid="stSidebar"] [data-testid="stFileUploader"] [class*="fileName"]{
  color: #f1f5f9 !important;
}

/* ── DataFrames ── */
[data-testid="stDataFrame"]{
  border-radius:12px; overflow:hidden;
  border:1px solid #e2e8f0;
  box-shadow: 0 4px 14px rgba(2,6,23,0.05);
}

/* ── Main area inputs ── */
.stTextInput > div > div > input,
.stNumberInput input,
.stDateInput input{
  border-radius:10px !important;
  border: 1px solid #cbd5e1 !important;
  background: #ffffff !important;
  color: #0f172a !important;
}
.stSelectbox > div > div{border-radius:10px !important;}
/* ── Sidebar selectbox text fix ── */
[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] > div,
[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] span,
[data-testid="stSidebar"] .stSelectbox [class*="placeholder"],
[data-testid="stSidebar"] .stSelectbox [class*="singleValue"]{
  color: #f1f5f9 !important;
}
[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] > div{
  background: #1e293b !important;
  border-color: #334155 !important;
}
.stButton > button{
  border-radius:12px !important;
  font-weight:700 !important;
}
</style>
""",
    unsafe_allow_html=True,
)


# ═══════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════
with st.sidebar:
    # ── Session header ──
    st.markdown(
        f"<p style='font-size:1.15rem;font-weight:700;color:#f1f5f9;margin:6px 0 4px 0;'>{st.session_state.username}</p>",
        unsafe_allow_html=True,
    )
    if st.button("Logout", use_container_width=True):
        st.session_state.update(
            authenticated=False, user_id=None,
            username=None, is_admin=False, must_change_pw=False,
        )
        st.rerun()
    st.divider()

    # ── MT5 Connection ──
    st.markdown("## MT5 Connection")

    # Saved configs selector
    _profiles = get_mt5_profiles(st.session_state.user_id)
    _profile_names = ["-- Enter manually --"] + [p["name"] for p in _profiles]
    _sel_name = st.selectbox("Saved Configs", _profile_names)

    _prefill_server   = ""
    _prefill_login    = 0
    _prefill_password = ""
    if _sel_name != "-- Enter manually --":
        _matched = next((p for p in _profiles if p["name"] == _sel_name), None)
        if _matched:
            _resolved = get_mt5_profile_decrypted(_matched["id"], st.session_state.user_id)
            if _resolved:
                _prefill_server   = _resolved["server"]
                _prefill_login    = _resolved["mt5_login"]
                _prefill_password = _resolved["mt5_password"]

    st.divider()

    server = st.text_input(
        "Server (IP:Port)",
        value=_prefill_server,
        placeholder="e.g. 188.240.63.240:443",
        help="MT5 server address with port",
    )
    login = st.number_input(
        "Manager Login",
        min_value=0,
        step=1,
        value=_prefill_login,
        help="Your MT5 Manager login ID",
    )
    password = st.text_input(
        "Password",
        value=_prefill_password,
        type="password",
        help="MT5 Manager password",
    )

    # ── Save / manage config ──
    with st.expander("Save as Config"):
        _save_name = st.text_input("Config name", placeholder="e.g. Main Broker", key="save_cfg_name")
        if st.button("Save", key="btn_save_cfg", use_container_width=True):
            if not _save_name.strip():
                st.error("Enter a config name.")
            elif not server or login == 0 or not password:
                st.error("Fill in all credentials before saving.")
            else:
                save_mt5_profile(st.session_state.user_id, _save_name.strip(), server, int(login), password)
                st.success(f"Saved as '{_save_name.strip()}'")
                st.rerun()

    if _sel_name != "-- Enter manually --":
        _matched = next((p for p in _profiles if p["name"] == _sel_name), None)
        if _matched:
            with st.expander("Manage Config"):
                st.caption(f"Selected: **{_sel_name}**")
                if st.button("Delete this config", key="btn_del_cfg", use_container_width=True):
                    delete_mt5_profile(_matched["id"], st.session_state.user_id)
                    st.rerun()

    st.divider()
    st.markdown("### Report Type")
    report_type = st.radio(
        "Select report",
        ["Deal P&L", "Equity P&L"],
        label_visibility="collapsed",
    )

    st.divider()

    if report_type == "Deal P&L":
        st.markdown("### Date Range")
        from_date = st.date_input(
            "From Date",
            value=datetime.utcnow().date() - timedelta(days=1),
        )
        to_date = st.date_input(
            "To Date",
            value=datetime.utcnow().date(),
        )
        st.divider()
        st.markdown("### Options")
        top_n = st.slider("Top N accounts", 5, 50, 10)
        # Equity placeholders (not used)
        oe_date = ce_date = None
        account_file = None

    else:  # Equity P&L
        st.markdown("### Equity Dates")
        oe_date = st.date_input(
            "Opening Equity Date",
            value=datetime.utcnow().date() - timedelta(days=7),
            help="End-of-day equity snapshot for this date is used as opening equity.",
        )
        ce_date = st.date_input(
            "Closing Equity Date",
            value=datetime.utcnow().date(),
            help="End-of-day equity snapshot for this date is used as closing equity.",
        )
        st.divider()
        st.markdown("### Account Filter *(optional)*")
        account_file = st.file_uploader(
            "Upload account list",
            type=["csv", "xlsx", "xls"],
            help="CSV or Excel with a 'Login' column (or logins in the first column). "
                 "If omitted, all accounts are included.",
        )
        st.divider()
        top_n = st.slider("Top N accounts", 5, 50, 10)
        # Deal placeholders (not used)
        from_date = to_date = None

    st.divider()
    generate = st.button("Generate Report", use_container_width=True)


# ═══════════════════════════════════════════════════════════
# HERO HEADER
# ═══════════════════════════════════════════════════════════
if report_type == "Deal P&L":
    hero_sub = (
        "Connect to MT5 → fetch closed deals → deal-based P&amp;L report.<br>"
        "<b>Formula:</b> Net PnL = Profit + Commission + Swap &nbsp;|&nbsp; "
        "Hit Ratio = Wins / (Wins + Losses) × 100"
    )
else:
    hero_sub = (
        "Connect to MT5 → fetch daily equity snapshots → equity-based P&amp;L.<br>"
        "<b>Formula:</b> Net P&amp;L = Closing Equity − Opening Equity − Net D/W − Net Credit − Bonus"
    )

st.markdown(
    f"""
<div class="hero">
  <div class="hero-badge">MT5 Manager API &nbsp;|&nbsp; {report_type}</div>
  <div class="hero-title">Reporting Tool</div>
  <p class="hero-sub">{hero_sub}</p>
</div>
""",
    unsafe_allow_html=True,
)

# ═══════════════════════════════════════════════════════════
# ADMIN PANEL
# ═══════════════════════════════════════════════════════════
if st.session_state.is_admin:
    with st.expander("Admin — User Management", expanded=False):
        st.markdown("#### Users")
        _users = list_users()
        for _u in _users:
            _col1, _col2, _col3 = st.columns([3, 1, 1])
            with _col1:
                _badge = "Admin" if _u["is_admin"] else "User"
                st.markdown(f"**{_u['username']}** &nbsp; `{_badge}`")
            with _col2:
                st.caption(_u["created_at"][:10])
            with _col3:
                _is_self = _u["id"] == st.session_state.user_id
                _admin_count = sum(1 for x in _users if x["is_admin"])
                _last_admin  = _u["is_admin"] and _admin_count == 1
                if not _is_self and not _last_admin:
                    if st.button("Delete", key=f"del_user_{_u['id']}", use_container_width=True):
                        delete_user(_u["id"])
                        st.rerun()
                else:
                    st.caption("—")

        st.divider()
        st.markdown("#### Create User")
        with st.form("create_user_form"):
            _new_uname  = st.text_input("Username")
            _new_pw     = st.text_input("Password", type="password")
            _new_admin  = st.checkbox("Admin")
            _new_change = st.checkbox("Force password change on first login", value=True)
            _created    = st.form_submit_button("Create", use_container_width=True)
        if _created:
            if not _new_uname.strip() or not _new_pw:
                st.error("Username and password are required.")
            else:
                try:
                    create_user(_new_uname.strip(), _new_pw, is_admin=int(_new_admin))
                    if _new_change:
                        from db import _conn, _hash_pw
                        with _conn() as _c:
                            _c.execute(
                                "UPDATE app_users SET must_change_pw=1 WHERE username=?",
                                (_new_uname.strip(),),
                            )
                    st.success(f"User '{_new_uname.strip()}' created.")
                    st.rerun()
                except ValueError as _e:
                    st.error(str(_e))


# ═══════════════════════════════════════════════════════════
# HELPER: load account filter from uploaded file
# ═══════════════════════════════════════════════════════════
def _load_account_filter(uploaded_file) -> list:
    """Parse uploaded CSV/Excel and return list of login integers."""
    if uploaded_file is None:
        return None
    try:
        if uploaded_file.name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(uploaded_file)
        else:
            df = pd.read_csv(uploaded_file)

        # Accept column named Login / login / ACCOUNT / account, else first column
        col = None
        for candidate in ("Login", "login", "ACCOUNT", "account", "AccountID"):
            if candidate in df.columns:
                col = candidate
                break
        if col is None:
            col = df.columns[0]

        logins = pd.to_numeric(df[col], errors="coerce").dropna().astype(int).tolist()
        return logins if logins else None
    except Exception as e:
        st.warning(f"Could not parse account file: {e}")
        return None


def _build_lp_excel(lp_calc: pd.DataFrame, all_clients_pnl):
    """Build a styled openpyxl workbook matching the B-Book report format."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    NAVY   = PatternFill("solid", fgColor="0B1220")
    MID    = PatternFill("solid", fgColor="1E3A5F")
    W_BOLD = Font(color="FFFFFF", bold=True, size=11)
    BOLD   = Font(bold=True)
    CTR    = Alignment(horizontal="center")
    NUM_FMT = "#,##0.00"

    wb = Workbook()
    ws = wb.active
    ws.title = "LP & Brokers PNL"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14

    # ── LP Summary blocks (columns B-C, one block per LP) ──
    start_row = 3
    for _, row in lp_calc.iterrows():
        oe        = float(row["Opening Equity"])
        ce        = float(row["Closing Equity"])
        diff      = float(row["Difference"])
        deposit   = float(row["Deposit"])
        credit_in = float(row["Credit IN"])
        additional= float(row["Additional"])
        withdrawal= float(row["Withdrawal"])
        credit_out= float(row["Credit OUT"])
        net_brok  = float(row["Net Brokerage"])

        r = start_row

        ws.merge_cells(f"B{r}:C{r}")
        ws[f"B{r}"].value = "LP Summery"
        ws[f"B{r}"].fill  = MID
        ws[f"B{r}"].font  = W_BOLD
        ws[f"B{r}"].alignment = CTR

        ws.merge_cells(f"B{r+1}:C{r+1}")
        ws[f"B{r+1}"].value = row.get("LP Name") or "LP"
        ws[f"B{r+1}"].fill  = NAVY
        ws[f"B{r+1}"].font  = W_BOLD
        ws[f"B{r+1}"].alignment = CTR

        fields = [
            ("Opening Equity", oe,         False),
            ("Closing Equity", ce,         False),
            ("Difference",     diff,       True),
            ("Deposit",        deposit,    False),
            ("Credit IN",      credit_in,  False),
            ("Additional",     additional, False),
            ("Withdrawal",     withdrawal, False),
            ("Credit OUT",     credit_out, False),
            ("Net Brokerage",  net_brok,   True),
        ]
        for i, (label, val, bold) in enumerate(fields):
            rr = r + 2 + i
            lc = ws[f"B{rr}"]
            vc = ws[f"C{rr}"]
            lc.value = label
            vc.value = round(val, 2)
            vc.number_format = NUM_FMT
            if bold:
                lc.font = BOLD
                vc.font = BOLD

        start_row = r + 2 + len(fields) + 2

    # ── Total Brokers PNL table (columns E-F, starting at row 3) ──
    total_lp = lp_calc["Net Brokerage"].sum()

    r = 3
    ws.merge_cells(f"E{r}:F{r}")
    ws[f"E{r}"].value = "Total Brokers PNL"
    ws[f"E{r}"].fill  = NAVY
    ws[f"E{r}"].font  = W_BOLD
    ws[f"E{r}"].alignment = CTR

    r += 1
    for col, hdr in [("E", "LP Name"), ("F", "PNL")]:
        ws[f"{col}{r}"].value = hdr
        ws[f"{col}{r}"].fill  = NAVY
        ws[f"{col}{r}"].font  = W_BOLD

    r += 1
    for _, row in lp_calc.iterrows():
        ws[f"E{r}"].value = row.get("LP Name") or "—"
        ws[f"F{r}"].value = round(float(row["Net Brokerage"]), 2)
        ws[f"F{r}"].number_format = NUM_FMT
        r += 1

    r += 1  # blank row
    ws[f"E{r}"].value = "Total"
    ws[f"E{r}"].font  = BOLD
    ws[f"F{r}"].value = round(total_lp, 2)
    ws[f"F{r}"].font  = BOLD
    ws[f"F{r}"].number_format = NUM_FMT

    r += 2  # blank row
    ws[f"E{r}"].value = "All Clients PNL"
    if all_clients_pnl is not None:
        ws[f"F{r}"].value = round(all_clients_pnl, 2)
        ws[f"F{r}"].number_format = NUM_FMT
    else:
        ws[f"F{r}"].value = "N/A"

    r += 2  # blank row
    net_total = round(total_lp - (all_clients_pnl or 0), 2)
    ws[f"E{r}"].value = "Net Brokerage"
    ws[f"E{r}"].font  = BOLD
    ws[f"F{r}"].value = net_total
    ws[f"F{r}"].fill  = NAVY
    ws[f"F{r}"].font  = W_BOLD
    ws[f"F{r}"].number_format = NUM_FMT

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════
# GENERATE REPORT
# ═══════════════════════════════════════════════════════════
if generate:

    # ── Shared validation ──
    if not server:
        st.error("Please enter the MT5 Server address (IP:Port).")
        st.stop()
    if login == 0:
        st.error("Please enter a valid Manager Login.")
        st.stop()
    if not password:
        st.error("Please enter the Manager Password.")
        st.stop()

    # ── Connect ──
    with st.spinner("Connecting to MT5 Manager API…"):
        manager, err = connect_mt5(server, int(login), password)
    if err:
        st.error(f"❌ {err}")
        st.stop()
    st.success("Connected to MT5 Manager API")

    # ════════════════════════════════════════════════════════
    # DEAL P&L REPORT
    # ════════════════════════════════════════════════════════
    if report_type == "Deal P&L":

        if from_date > to_date:
            st.error("'From Date' must be before 'To Date'.")
            disconnect_mt5(manager)
            st.stop()

        dt_from = datetime.combine(from_date, datetime.min.time())
        dt_to   = datetime.combine(to_date,   datetime.max.time())

        try:
            with st.spinner(f"Fetching deals {from_date} → {to_date}…"):
                deals_raw, err = get_deals(manager, dt_from, dt_to)
            if err:
                st.warning(f"⚠️ {err}")

            if not deals_raw:
                st.warning("No trade deals found in the selected date range.")
                disconnect_mt5(manager)
                st.stop()

            deal_dicts  = deals_to_dicts(deals_raw, {})
            deal_logins = list({d["Login"] for d in deal_dicts if d.get("Login")})

            with st.spinner(f"Fetching group data for {len(deal_logins):,} logins…"):
                user_map, err = get_users(manager, deal_logins)
            if err:
                st.warning(f"⚠️ {err}")

            for d in deal_dicts:
                d["Group"] = user_map.get(d["Login"], "")

            df_deals = build_deals_dataframe(deal_dicts)

            with st.spinner("Crunching numbers…"):
                report     = compute_client_report(df_deals)
                group_df   = compute_group_summary(report)
                symbol_df  = compute_symbol_summary(df_deals)
                kpis       = compute_kpis(report)

            # ── KPI Overview ──
            st.markdown(
                """<div class="section"><div class="section-title">
                <h2>Overview</h2><span class="pill">Deal P&L KPIs</span>
                </div></div>""",
                unsafe_allow_html=True,
            )

            pnl_color = "green" if kpis["total_pnl"] >= 0 else "red"
            k1, k2, k3, k4 = st.columns(4)
            with k1:
                st.markdown(
                    f'<div class="metric"><div class="k">Clients</div>'
                    f'<div class="v">{kpis["total_clients"]:,}</div>'
                    f'<div class="s">Unique logins with trades</div></div>',
                    unsafe_allow_html=True,
                )
            with k2:
                st.markdown(
                    f'<div class="metric"><div class="k">Net Client PnL</div>'
                    f'<div class="v {pnl_color}">${kpis["total_pnl"]:,.2f}</div>'
                    f'<div class="s">Profit + Commission + Swap</div></div>',
                    unsafe_allow_html=True,
                )
            with k3:
                st.markdown(
                    f'<div class="metric"><div class="k">Closed Lots</div>'
                    f'<div class="v">{kpis["total_lots"]:,.2f}</div>'
                    f'<div class="s">Sum of deal volumes</div></div>',
                    unsafe_allow_html=True,
                )
            with k4:
                st.markdown(
                    f'<div class="metric"><div class="k">Total Trades</div>'
                    f'<div class="v">{kpis["total_trades"]:,}</div>'
                    f'<div class="s">Avg hit ratio: {kpis["avg_hit_ratio"]:.1f}%</div></div>',
                    unsafe_allow_html=True,
                )

            k5, k6, k7, k8 = st.columns(4)
            with k5:
                st.markdown(
                    f'<div class="metric"><div class="k">Total Profit</div>'
                    f'<div class="v green">${kpis["total_profit"]:,.2f}</div>'
                    f'<div class="s">Accounts with PnL &gt; 0</div></div>',
                    unsafe_allow_html=True,
                )
            with k6:
                st.markdown(
                    f'<div class="metric"><div class="k">Total Loss</div>'
                    f'<div class="v red">${kpis["total_loss"]:,.2f}</div>'
                    f'<div class="s">Accounts with PnL &lt; 0</div></div>',
                    unsafe_allow_html=True,
                )
            with k7:
                st.markdown(
                    f'<div class="metric"><div class="k">Volume USD</div>'
                    f'<div class="v">${kpis["total_volume"]:,.0f}</div>'
                    f'<div class="s">Lots × 100,000</div></div>',
                    unsafe_allow_html=True,
                )
            with k8:
                st.markdown(
                    f'<div class="metric"><div class="k">Date Range</div>'
                    f'<div class="v" style="font-size:1.1rem;">{from_date} → {to_date}</div>'
                    f'<div class="s">Report period</div></div>',
                    unsafe_allow_html=True,
                )

            chart_data = pd.DataFrame(
                {"Side": ["Profit", "Loss"], "Amount": [kpis["total_profit"], abs(kpis["total_loss"])]}
            ).set_index("Side")
            st.bar_chart(chart_data, height=240)

            # ── Top Gainers / Losers ──
            st.markdown(
                f"""<div class="section"><div class="section-title">
                <h2>Top {top_n} Accounts</h2><span class="pill">Leaders</span>
                </div></div>""",
                unsafe_allow_html=True,
            )

            display_cols = ["Login", "Group", "Closed Lots", "NET PNL USD", "Total Trades", "Wins", "Losses", "Hit Ratio %"]
            t1, t2 = st.columns(2)
            with t1:
                st.markdown(f"**Top {top_n} Gainers**")
                st.dataframe(
                    get_top_gainers(report, top_n)[display_cols].style.format(
                        {"NET PNL USD": "${:,.2f}", "Closed Lots": "{:,.2f}", "Hit Ratio %": "{:.1f}%"}
                    ),
                    use_container_width=True,
                )
            with t2:
                st.markdown(f"**Top {top_n} Losers**")
                st.dataframe(
                    get_top_losers(report, top_n)[display_cols].style.format(
                        {"NET PNL USD": "${:,.2f}", "Closed Lots": "{:,.2f}", "Hit Ratio %": "{:.1f}%"}
                    ),
                    use_container_width=True,
                )

            # ── Group Summary ──
            if not group_df.empty:
                st.markdown(
                    """<div class="section"><div class="section-title">
                    <h2>Group Summary</h2><span class="pill">Aggregated</span>
                    </div></div>""",
                    unsafe_allow_html=True,
                )
                g1, g2 = st.columns(2)
                with g1:
                    st.markdown("**Top profit groups**")
                    st.dataframe(group_df.sort_values("NET_PNL_USD", ascending=False).head(top_n), use_container_width=True)
                with g2:
                    st.markdown("**Top loss groups**")
                    st.dataframe(group_df.sort_values("NET_PNL_USD", ascending=True).head(top_n), use_container_width=True)

            # ── Symbol Breakdown ──
            if not symbol_df.empty:
                st.markdown(
                    """<div class="section"><div class="section-title">
                    <h2>Symbol Breakdown</h2><span class="pill">Instruments</span>
                    </div></div>""",
                    unsafe_allow_html=True,
                )
                st.dataframe(
                    symbol_df.head(30).style.format({"NET_PNL_USD": "${:,.2f}", "Closed_Lots": "{:,.2f}"}),
                    use_container_width=True,
                )

            # ── Full Client Table ──
            st.markdown(
                """<div class="section"><div class="section-title">
                <h2>Full Client Report</h2><span class="pill">All Accounts</span>
                </div></div>""",
                unsafe_allow_html=True,
            )
            all_groups = ["All"] + sorted(report["Group"].dropna().unique().tolist())
            selected_group = st.selectbox("Filter by Group", all_groups)
            filtered = report if selected_group == "All" else report[report["Group"] == selected_group]
            st.dataframe(
                filtered.style.format({
                    "NET PNL USD": "${:,.2f}", "Volume USD": "${:,.0f}",
                    "Closed Lots": "{:,.2f}", "Hit Ratio %": "{:.1f}%",
                    "Commission": "${:,.2f}", "Swap": "${:,.2f}",
                }),
                use_container_width=True,
            )
            st.caption(f"Showing {len(filtered):,} of {len(report):,} accounts")

            # ── Export ──
            st.markdown(
                """<div class="section"><div class="section-title">
                <h2>Export</h2><span class="pill">Download</span>
                </div></div>""",
                unsafe_allow_html=True,
            )
            exp1, exp2 = st.columns(2)
            with exp1:
                st.download_button(
                    "Download CSV",
                    data=report.to_csv(index=False),
                    file_name=f"DealPnL_{from_date}_to_{to_date}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with exp2:
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as w:
                    report.to_excel(w, index=False, sheet_name="Client PnL")
                    group_df.to_excel(w, index=False, sheet_name="Group Summary")
                    symbol_df.to_excel(w, index=False, sheet_name="Symbol Breakdown")
                buf.seek(0)
                st.download_button(
                    "Download Excel",
                    data=buf,
                    file_name=f"DealPnL_{from_date}_to_{to_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            # ── Download All ──
            kpi_rows = [
                ("Report Type",    "Deal P&L"),
                ("Date Range",     f"{from_date} → {to_date}"),
                ("Total Clients",  kpis["total_clients"]),
                ("Net Client PnL", f"${kpis['total_pnl']:,.2f}"),
                ("Total Profit",   f"${kpis['total_profit']:,.2f}"),
                ("Total Loss",     f"${kpis['total_loss']:,.2f}"),
                ("Closed Lots",    f"{kpis['total_lots']:,.2f}"),
                ("Volume USD",     f"${kpis['total_volume']:,.0f}"),
                ("Total Trades",   kpis["total_trades"]),
                ("Avg Hit Ratio",  f"{kpis['avg_hit_ratio']:.1f}%"),
            ]
            kpi_summary = pd.DataFrame(kpi_rows, columns=["Metric", "Value"])

            buf_all = BytesIO()
            with pd.ExcelWriter(buf_all, engine="openpyxl") as w:
                kpi_summary.to_excel(w, index=False, sheet_name="KPI Summary")
                get_top_gainers(report, 50).to_excel(w, index=False, sheet_name="Top 50 Gainers")
                get_top_losers(report, 50).to_excel(w, index=False, sheet_name="Top 50 Losers")
                report.to_excel(w, index=False, sheet_name="Full Client Report")
                if not group_df.empty:
                    group_df.to_excel(w, index=False, sheet_name="Group Summary")
                if not symbol_df.empty:
                    symbol_df.to_excel(w, index=False, sheet_name="Symbol Breakdown")
            buf_all.seek(0)
            st.download_button(
                "Download All — Full Report (Excel)",
                data=buf_all,
                file_name=f"DealPnL_FULL_{from_date}_to_{to_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except Exception as e:
            st.error(f"❌ Error generating report: {e}")
            import traceback
            st.code(traceback.format_exc())
        finally:
            disconnect_mt5(manager)
            st.info("MT5 connection closed.")

    # ════════════════════════════════════════════════════════
    # EQUITY P&L REPORT
    # ════════════════════════════════════════════════════════
    else:
        if oe_date > ce_date:
            st.error("Opening Equity Date must be before (or equal to) Closing Equity Date.")
            disconnect_mt5(manager)
            st.stop()

        account_filter = _load_account_filter(account_file)
        if account_filter is not None:
            st.info(f"Account filter loaded: {len(account_filter):,} logins")

        try:
            # Date boundaries
            oe_dt_from  = datetime.combine(oe_date, datetime.min.time())
            oe_dt_to    = datetime.combine(oe_date, datetime.max.time())
            ce_dt_from  = datetime.combine(ce_date, datetime.min.time())
            ce_dt_to    = datetime.combine(ce_date, datetime.max.time())

            # Summary period: day after OE through CE (captures all D/W between snapshots)
            sum_dt_from = datetime.combine(oe_date + timedelta(days=1), datetime.min.time())
            sum_dt_to   = datetime.combine(ce_date, datetime.max.time())
            summary_available = sum_dt_from <= sum_dt_to

            # ── Fetch Opening Equity snapshots ──
            with st.spinner(f"Fetching Opening Equity snapshots ({oe_date})…"):
                oe_raw, err = get_daily_reports(manager, oe_dt_from, oe_dt_to)
            if err:
                st.warning(f"⚠️ Opening Equity fetch: {err}")
            oe_dicts = daily_to_dicts(oe_raw) if oe_raw else []

            # ── Fetch Closing Equity snapshots ──
            with st.spinner(f"Fetching Closing Equity snapshots ({ce_date})…"):
                ce_raw, err = get_daily_reports(manager, ce_dt_from, ce_dt_to)
            if err:
                st.warning(f"⚠️ Closing Equity fetch: {err}")
            ce_dicts = daily_to_dicts(ce_raw) if ce_raw else []

            # ── Fetch summary period (D/W, credit, bonus) ──
            summary_dicts = []
            if summary_available:
                with st.spinner(f"Fetching activity {oe_date + timedelta(days=1)} → {ce_date}…"):
                    sm_raw, err = get_daily_reports(manager, sum_dt_from, sum_dt_to)
                if err:
                    st.warning(f"⚠️ Summary period fetch: {err}")
                summary_dicts = daily_to_dicts(sm_raw) if sm_raw else []

            if not oe_dicts and not ce_dicts:
                st.warning("No daily equity data found for the selected dates. "
                           "Ensure the MT5 server has Daily Reports enabled and the dates are valid.")
                disconnect_mt5(manager)
                st.stop()

            # ── Fetch group data ──
            data_logins = set()
            for dicts in (oe_dicts, ce_dicts, summary_dicts):
                data_logins.update(d["Login"] for d in dicts)
            if account_filter:
                missed_logins = sorted(set(account_filter) - data_logins)
                all_logins = data_logins.intersection(set(account_filter))
            else:
                missed_logins = []
                all_logins = data_logins

            logins_list = list(all_logins)
            with st.spinner(f"Fetching group data for {len(logins_list):,} logins…"):
                user_map, err = get_users(manager, logins_list)
            if err:
                st.warning(f"⚠️ {err}")

            # ── Compute equity report ──
            with st.spinner("Computing equity P&L…"):
                eq_report  = compute_equity_report(
                    oe_dicts, ce_dicts, summary_dicts, user_map, account_filter
                )
                eq_groups  = compute_equity_group_summary(eq_report)
                eq_kpis    = compute_equity_kpis(eq_report)
                st.session_state["eq_total_net_pnl"] = eq_kpis["total_net_pnl"]

            # ── Fetch registration dates ──
            with st.spinner("Fetching account registration data…"):
                _reg_map, _reg_err = get_user_registrations(manager, logins_list)
            if _reg_err:
                st.warning(f"⚠️ Registration fetch: {_reg_err}")
            _oe_ts = int(datetime.combine(oe_date, datetime.min.time()).timestamp())
            _ce_ts = int(datetime.combine(ce_date, datetime.max.time()).timestamp())
            _new_reg_count = sum(
                1 for ts in _reg_map.values() if ts and _oe_ts <= ts <= _ce_ts
            )
            _acct_summary = compute_account_category_summary(eq_report, _new_reg_count)

            if eq_report.empty:
                st.warning("No equity data found for the selected accounts and dates.")
                disconnect_mt5(manager)
                st.stop()

            # ── Missed accounts warning ──
            if missed_logins:
                with st.expander(f"⚠️ {len(missed_logins):,} account(s) from your filter had no snapshot data for the selected dates"):
                    missed_df = pd.DataFrame({"Login (not in report)": missed_logins})
                    st.dataframe(missed_df, use_container_width=True, hide_index=True)

            # ════════════════════════════════
            # KPI OVERVIEW
            # ════════════════════════════════
            st.markdown(
                """<div class="section"><div class="section-title">
                <h2>Overview</h2><span class="pill">Equity P&L KPIs</span>
                </div></div>""",
                unsafe_allow_html=True,
            )

            pnl_color = "green" if eq_kpis["total_net_pnl"] >= 0 else "red"
            k1, k2, k3, k4 = st.columns(4)
            with k1:
                st.markdown(
                    f'<div class="metric"><div class="k">Accounts</div>'
                    f'<div class="v">{eq_kpis["total_accounts"]:,}</div>'
                    f'<div class="s">In report</div></div>',
                    unsafe_allow_html=True,
                )
            with k2:
                st.markdown(
                    f'<div class="metric"><div class="k">Net P&L</div>'
                    f'<div class="v {pnl_color}">${eq_kpis["total_net_pnl"]:,.2f}</div>'
                    f'<div class="s">CE − OE − D/W − Credit − Bonus</div></div>',
                    unsafe_allow_html=True,
                )
            with k3:
                st.markdown(
                    f'<div class="metric"><div class="k">Opening Equity</div>'
                    f'<div class="v blue">${eq_kpis["total_oe"]:,.2f}</div>'
                    f'<div class="s">{oe_date}</div></div>',
                    unsafe_allow_html=True,
                )
            with k4:
                st.markdown(
                    f'<div class="metric"><div class="k">Closing Equity</div>'
                    f'<div class="v blue">${eq_kpis["total_ce"]:,.2f}</div>'
                    f'<div class="s">{ce_date}</div></div>',
                    unsafe_allow_html=True,
                )

            k5, k6, k7, k8 = st.columns(4)
            with k5:
                st.markdown(
                    f'<div class="metric"><div class="k">Net D/W</div>'
                    f'<div class="v">${eq_kpis["total_net_dw"]:,.2f}</div>'
                    f'<div class="s">Deposits + Withdrawals</div></div>',
                    unsafe_allow_html=True,
                )
            with k6:
                st.markdown(
                    f'<div class="metric"><div class="k">Net Credit</div>'
                    f'<div class="v">${eq_kpis["total_net_credit"]:,.2f}</div>'
                    f'<div class="s">Credit In + Credit Out</div></div>',
                    unsafe_allow_html=True,
                )
            with k7:
                st.markdown(
                    f'<div class="metric"><div class="k">Bonus</div>'
                    f'<div class="v">${eq_kpis["total_bonus"]:,.2f}</div>'
                    f'<div class="s">Total bonus added</div></div>',
                    unsafe_allow_html=True,
                )
            with k8:
                st.markdown(
                    f'<div class="metric"><div class="k">Profitable / Losing</div>'
                    f'<div class="v" style="font-size:1.2rem;">'
                    f'<span style="color:#059669">{eq_kpis["profitable_accounts"]}</span>'
                    f' / <span style="color:#dc2626">{eq_kpis["losing_accounts"]}</span></div>'
                    f'<div class="s">Accounts by Net P&amp;L sign</div></div>',
                    unsafe_allow_html=True,
                )

            # Equity change chart
            chart_data = pd.DataFrame({
                "Metric": ["Opening Equity", "Closing Equity"],
                "Value":  [eq_kpis["total_oe"], eq_kpis["total_ce"]],
            }).set_index("Metric")
            st.bar_chart(chart_data, height=240)

            # ════════════════════════════════
            # TOP GAINERS / LOSERS
            # ════════════════════════════════
            st.markdown(
                f"""<div class="section"><div class="section-title">
                <h2>Top {top_n} Accounts</h2><span class="pill">Leaders</span>
                </div></div>""",
                unsafe_allow_html=True,
            )

            eq_display = ["Login", "Group", "Opening Equity", "Closing Equity", "Net D/W", "Net Credit", "Bonus", "Net P&L"]
            eq_fmt = {c: "${:,.2f}" for c in ["Opening Equity", "Closing Equity", "Net D/W", "Net Credit", "Bonus", "Net P&L"]}

            t1, t2 = st.columns(2)
            with t1:
                st.markdown(f"**Top {top_n} Gainers**")
                st.dataframe(
                    eq_report.sort_values("Net P&L", ascending=False).head(top_n)[eq_display].style.format(eq_fmt),
                    use_container_width=True,
                )
            with t2:
                st.markdown(f"**Top {top_n} Losers**")
                st.dataframe(
                    eq_report.sort_values("Net P&L", ascending=True).head(top_n)[eq_display].style.format(eq_fmt),
                    use_container_width=True,
                )

            # ════════════════════════════════
            # GROUP SUMMARY
            # ════════════════════════════════
            if not eq_groups.empty:
                st.markdown(
                    """<div class="section"><div class="section-title">
                    <h2>Group Summary</h2><span class="pill">Aggregated</span>
                    </div></div>""",
                    unsafe_allow_html=True,
                )
                st.dataframe(
                    eq_groups.style.format({
                        "Opening_Equity": "${:,.2f}", "Closing_Equity": "${:,.2f}",
                        "Net_DW": "${:,.2f}", "Net_Credit": "${:,.2f}",
                        "Bonus": "${:,.2f}", "Net_PnL": "${:,.2f}",
                    }),
                    use_container_width=True,
                )

            # ════════════════════════════════
            # ACCOUNT CATEGORY SUMMARY
            # ════════════════════════════════
            st.markdown(
                """<div class="section"><div class="section-title">
                <h2>Account Summary</h2><span class="pill">Equity Categories</span>
                </div></div>""",
                unsafe_allow_html=True,
            )
            st.dataframe(_acct_summary, use_container_width=True, hide_index=True)

            # ════════════════════════════════
            # FULL ACCOUNT TABLE
            # ════════════════════════════════
            st.markdown(
                """<div class="section"><div class="section-title">
                <h2>Full Account Report</h2><span class="pill">All Accounts</span>
                </div></div>""",
                unsafe_allow_html=True,
            )

            money_cols = ["Opening Equity", "Closing Equity", "Difference",
                          "Deposits", "Withdrawals", "Net D/W",
                          "Credit In", "Credit Out", "Net Credit", "Bonus", "Net P&L"]
            fmt = {c: "${:,.2f}" for c in money_cols}

            all_groups_list = ["All"] + sorted(eq_report["Group"].dropna().unique().tolist())
            sel_group = st.selectbox("Filter by Group", all_groups_list)
            filtered_eq = eq_report if sel_group == "All" else eq_report[eq_report["Group"] == sel_group]

            st.dataframe(filtered_eq.style.format(fmt), use_container_width=True)
            st.caption(f"Showing {len(filtered_eq):,} of {len(eq_report):,} accounts")

            # ════════════════════════════════
            # EXPORT
            # ════════════════════════════════
            st.markdown(
                """<div class="section"><div class="section-title">
                <h2>Export</h2><span class="pill">Download</span>
                </div></div>""",
                unsafe_allow_html=True,
            )
            exp1, exp2 = st.columns(2)
            with exp1:
                st.download_button(
                    "Download CSV",
                    data=eq_report.to_csv(index=False),
                    file_name=f"EquityPnL_{oe_date}_to_{ce_date}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with exp2:
                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as w:
                    eq_report.to_excel(w, index=False, sheet_name="Equity PnL")
                    if not eq_groups.empty:
                        eq_groups.to_excel(w, index=False, sheet_name="Group Summary")
                    _acct_summary.to_excel(w, index=False, sheet_name="Account Summary")
                    if missed_logins:
                        pd.DataFrame({"Login (Not Fetched)": missed_logins}).to_excel(
                            w, index=False, sheet_name="Not Fetched"
                        )
                buf.seek(0)
                st.download_button(
                    "Download Excel",
                    data=buf,
                    file_name=f"EquityPnL_{oe_date}_to_{ce_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            # ── Download All ──
            eq_kpi_rows = [
                ("Report Type",        "Equity P&L"),
                ("Opening Equity Date", str(oe_date)),
                ("Closing Equity Date", str(ce_date)),
                ("Total Accounts",     eq_kpis["total_accounts"]),
                ("Opening Equity",     f"${eq_kpis['total_oe']:,.2f}"),
                ("Closing Equity",     f"${eq_kpis['total_ce']:,.2f}"),
                ("Net D/W",            f"${eq_kpis['total_net_dw']:,.2f}"),
                ("Net Credit",         f"${eq_kpis['total_net_credit']:,.2f}"),
                ("Bonus",              f"${eq_kpis['total_bonus']:,.2f}"),
                ("Net P&L",            f"${eq_kpis['total_net_pnl']:,.2f}"),
                ("Profitable Accounts", eq_kpis["profitable_accounts"]),
                ("Losing Accounts",    eq_kpis["losing_accounts"]),
            ]
            eq_kpi_summary = pd.DataFrame(eq_kpi_rows, columns=["Metric", "Value"])

            buf_all = BytesIO()
            with pd.ExcelWriter(buf_all, engine="openpyxl") as w:
                eq_kpi_summary.to_excel(w, index=False, sheet_name="KPI Summary")
                eq_report.sort_values("Net P&L", ascending=False).head(50).to_excel(
                    w, index=False, sheet_name="Top 50 Gainers"
                )
                eq_report.sort_values("Net P&L", ascending=True).head(50).to_excel(
                    w, index=False, sheet_name="Top 50 Losers"
                )
                eq_report.to_excel(w, index=False, sheet_name="Full Account Report")
                if not eq_groups.empty:
                    eq_groups.to_excel(w, index=False, sheet_name="Group Summary")
                _acct_summary.to_excel(w, index=False, sheet_name="Account Summary")
                if missed_logins:
                    pd.DataFrame({"Login (Not Fetched)": missed_logins}).to_excel(
                        w, index=False, sheet_name="Not Fetched"
                    )
                pass  # LP sheets appended below via styled helper
            # If LP data exists, merge LP sheet into buf_all
            _lp_data = st.session_state.get("lp_df_computed")
            _lp_named_all = _lp_data[_lp_data["LP Name"].astype(str).str.strip() != ""] if _lp_data is not None else None
            if _lp_named_all is not None and not _lp_named_all.empty:
                from openpyxl import load_workbook
                _lp_wb_buf = _build_lp_excel(_lp_named_all, st.session_state.get("eq_total_net_pnl"))
                _lp_wb     = load_workbook(_lp_wb_buf)
                _main_wb   = load_workbook(buf_all)
                _lp_ws     = _lp_wb.active
                _new_ws    = _main_wb.create_sheet("LP & Brokers PNL")
                for row in _lp_ws.iter_rows():
                    for cell in row:
                        _nc = _new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            _nc.font       = cell.font.copy()
                            _nc.fill       = cell.fill.copy()
                            _nc.alignment  = cell.alignment.copy()
                            _nc.number_format = cell.number_format
                for col, dim in _lp_ws.column_dimensions.items():
                    _new_ws.column_dimensions[col].width = dim.width
                for rng in _lp_ws.merged_cells.ranges:
                    _new_ws.merge_cells(str(rng))
                buf_all = BytesIO()
                _main_wb.save(buf_all)
            buf_all.seek(0)
            st.download_button(
                "Download All — Full Report (Excel)",
                data=buf_all,
                file_name=f"EquityPnL_FULL_{oe_date}_to_{ce_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except Exception as e:
            st.error(f"❌ Error generating equity report: {e}")
            import traceback
            st.code(traceback.format_exc())
        finally:
            disconnect_mt5(manager)
            st.info("MT5 connection closed.")


# ═══════════════════════════════════════════════════════════
# DEFAULT STATE (before generate is clicked)
# ═══════════════════════════════════════════════════════════
else:
    if report_type == "Deal P&L":
        st.markdown(
            """
<div class="section"><div class="section-title">
<h2>Getting Started — Deal P&L</h2><span class="pill">Guide</span>
</div></div>
""",
            unsafe_allow_html=True,
        )
        st.markdown(
            """
**How it works:**

1. Enter MT5 Manager credentials in the sidebar
2. Select a **From / To date range** for closed deals
3. Click **Generate Report**
4. View KPIs, top gainers/losers, group and symbol breakdowns
5. Export as CSV or Excel

**Formula:** PnL = Profit + Commission + Swap
**Hit Ratio:** Wins / (Wins + Losses) × 100
**Volume USD:** Closed Lots × 100,000 (standard forex lot)
"""
        )
    else:
        st.markdown(
            """
<div class="section"><div class="section-title">
<h2>Getting Started — Equity P&L</h2><span class="pill-green">Guide</span>
</div></div>
""",
            unsafe_allow_html=True,
        )
        st.markdown(
            """
**How it works:**

1. Enter MT5 Manager credentials in the sidebar
2. Select an **Opening Equity Date** and **Closing Equity Date**
3. Optionally upload an **account list** (CSV/Excel with a Login column) to filter specific accounts
4. Click **Generate Report**
5. View equity-based P&L per account with full D/W and credit breakdown
6. Export as CSV or Excel

**Formula:**
`Net P&L = Closing Equity − Opening Equity − Net D/W − Net Credit − Bonus`

**Data source:** MT5 Daily Reports (`DailyRequestByGroupNumPy`)
- **Opening / Closing Equity** = `ProfitEquity` on the selected date (end-of-day equity incl. floating)
- **Net D/W** = sum of `DailyBalance` between OE+1 and CE date
- **Net Credit** = sum of `DailyCredit` between OE+1 and CE date
- **Bonus** = sum of `DailyBonus` between OE+1 and CE date

**Account filter:** If no file is uploaded, all accounts with daily data are included.
"""
        )


# ═══════════════════════════════════════════════════════════
# LP SUMMARY (B BOOK REPORT) — Equity P&L mode only
# ═══════════════════════════════════════════════════════════
if report_type == "Equity P&L":
    st.markdown(
        """<div class="section"><div class="section-title">
        <h2>LP Summary</h2><span class="pill">Manual Entry</span>
        </div></div>""",
        unsafe_allow_html=True,
    )

    _lp_num_cols = ["Opening Equity", "Closing Equity", "Deposit", "Credit IN",
                    "Additional", "Withdrawal", "Credit OUT"]
    _lp_blank = {"LP Name": "", **{c: 0.0 for c in _lp_num_cols}}

    if "lp_df" not in st.session_state:
        # Pre-populate 20 empty rows so users can paste data directly
        st.session_state["lp_df"] = pd.DataFrame([dict(_lp_blank) for _ in range(20)])

    st.caption(
        "Enter or paste LP equity data below (up to 20 LPs). "
        "Blank numeric cells default to 0. Net Brokerage is calculated automatically."
    )

    _edited = st.data_editor(
        st.session_state["lp_df"],
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "LP Name":        st.column_config.TextColumn("LP Name"),
            "Opening Equity": st.column_config.NumberColumn("Opening Equity", format="%.2f", default=0.0),
            "Closing Equity": st.column_config.NumberColumn("Closing Equity", format="%.2f", default=0.0),
            "Deposit":        st.column_config.NumberColumn("Deposit",        format="%.2f", default=0.0),
            "Credit IN":      st.column_config.NumberColumn("Credit IN",      format="%.2f", default=0.0),
            "Additional":     st.column_config.NumberColumn("Additional",     format="%.2f", default=0.0),
            "Withdrawal":     st.column_config.NumberColumn("Withdrawal",     format="%.2f", default=0.0),
            "Credit OUT":     st.column_config.NumberColumn("Credit OUT",     format="%.2f", default=0.0),
        },
        key="lp_editor",
    )
    # Fill any None/NaN in numeric columns with 0
    _edited[_lp_num_cols] = _edited[_lp_num_cols].fillna(0.0)
    st.session_state["lp_df"] = _edited

    # ── Compute derived columns ──
    _lp_calc = _edited.copy()
    _lp_calc["Difference"] = _lp_calc["Closing Equity"] - _lp_calc["Opening Equity"]
    _lp_calc["Net Brokerage"] = (
        _lp_calc["Difference"]
        - _lp_calc["Deposit"]
        - _lp_calc["Credit IN"]
        - _lp_calc["Additional"]
        + _lp_calc["Withdrawal"]
        + _lp_calc["Credit OUT"]
    )
    st.session_state["lp_df_computed"] = _lp_calc

    # ── Results: per-LP summary + Total Brokers PNL ──
    _total_lp_pnl    = _lp_calc["Net Brokerage"].sum()
    _all_clients_pnl = st.session_state.get("eq_total_net_pnl", None)

    # Per-LP cards
    _lp_named = _lp_calc[_lp_calc["LP Name"].astype(str).str.strip() != ""]
    if not _lp_named.empty:
        st.markdown("**LP Results**")
        _card_cols = st.columns(min(len(_lp_named), 3))
        for _ci, (_, _lrow) in enumerate(_lp_named.iterrows()):
            with _card_cols[_ci % len(_card_cols)]:
                _nb_color = "#059669" if _lrow["Net Brokerage"] >= 0 else "#dc2626"
                st.markdown(
                    f'<div class="metric">'
                    f'<div class="k">{_lrow["LP Name"]}</div>'
                    f'<div class="v" style="color:{_nb_color};">${_lrow["Net Brokerage"]:,.2f}</div>'
                    f'<div class="s">Diff: ${_lrow["Difference"]:,.2f} &nbsp;|&nbsp; '
                    f'OE: ${_lrow["Opening Equity"]:,.2f} &nbsp;|&nbsp; CE: ${_lrow["Closing Equity"]:,.2f}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

    # Total Brokers PNL table
    st.markdown("#### Total Brokers PNL")
    _brokers_rows = []
    for _, _row in _lp_named.iterrows():
        _brokers_rows.append({"Item": _row["LP Name"], "PNL": _row["Net Brokerage"]})
    _brokers_rows.append({"Item": "Total", "PNL": _total_lp_pnl})

    if _all_clients_pnl is not None:
        _brokers_rows.append({"Item": "All Clients PNL", "PNL": _all_clients_pnl})
        _net_brokerage = _total_lp_pnl - _all_clients_pnl
        _brokers_rows.append({"Item": "Net Brokerage",   "PNL": _net_brokerage})
        # Show as metrics
        _mb1, _mb2, _mb3 = st.columns(3)
        _nb_col = "green" if _net_brokerage >= 0 else "red"
        with _mb1:
            st.markdown(
                f'<div class="metric"><div class="k">Total LP PNL</div>'
                f'<div class="v">${_total_lp_pnl:,.2f}</div></div>',
                unsafe_allow_html=True,
            )
        with _mb2:
            st.markdown(
                f'<div class="metric"><div class="k">All Clients PNL</div>'
                f'<div class="v">${_all_clients_pnl:,.2f}</div></div>',
                unsafe_allow_html=True,
            )
        with _mb3:
            st.markdown(
                f'<div class="metric"><div class="k">Net Brokerage</div>'
                f'<div class="v {_nb_col}">${_net_brokerage:,.2f}</div>'
                f'<div class="s">Total LP PNL − All Clients PNL</div></div>',
                unsafe_allow_html=True,
            )
    else:
        _brokers_df_tmp = pd.DataFrame(_brokers_rows)
        st.dataframe(_brokers_df_tmp.style.format({"PNL": "${:,.2f}"}), use_container_width=True)
        st.info("Generate the Equity P&L report first to see All Clients PNL and Net Brokerage.")

    _brokers_df = pd.DataFrame(_brokers_rows)
    st.session_state["lp_brokers_pnl"] = _brokers_df

    # ── Standalone export (styled) ──
    if not _lp_named.empty:
        _lp_buf = _build_lp_excel(_lp_named, _all_clients_pnl)
        st.download_button(
            "Download LP Summary Excel",
            data=_lp_buf,
            file_name="LP_Summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
